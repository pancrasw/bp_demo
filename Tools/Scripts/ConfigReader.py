from distutils.log import debug
import openpyxl
import json
import os

# 注意，要先关闭excel表格，再调用该程序

input_dir_path = "C:/Users/LEGION/Documents/GitHub/bp_demo/Tools/Configs"
output_dir_path = "C:/Users/LEGION/Documents/GitHub/bp_demo/Assets/Scripts/json"


class ConfigReader:
    # name表示输出的json文件名称
    def __init__(self, path: str, name: str) -> None:
        self.wb = openpyxl.load_workbook(path)
        self.name = name
        print("sheetnames:", self.wb.sheetnames)
        pass

    def close(self):
        self.wb.close()

    def outputJson(self, path: str):
        dicts = []
        file_postfix = ".json"
        wb = self.wb
        for sheet_name in self.wb.sheetnames:
            dicts = self._createDict(wb[sheet_name])
            print(dicts)
            json_str = json.dumps(dicts)
            file_name = path+os.sep+self.name+'.' + \
                sheet_name+file_postfix  # 完成路径，格式为文件名.表格名.json
            f = open(file_name, "w")
            print("start write "+file_name)
            f.write(json_str)
            f.close()
        pass

    # 解析表格返回字典
    def _createDict(self, sheet) -> list[dict]:
        attributes = []
        all_result = []
        attribute_row = sheet[2]
        for cell in attribute_row:
            attributes.append(cell.value)
        for row_i in range(3, sheet.max_row+1):  # correct
            record_dict = {}  # 单条记录的字典
            record_row = sheet[row_i]  # 表格中的一行
            for i in range(0, len(attributes)):
                record_dict[attributes[i]] = record_row[i].value
            all_result.append(record_dict)
        return all_result


class ConfigManager:
    def __init__(self, dir_path):
        self.dirPath = dir_path
        self.fileNameList = os.listdir(dir_path)
        print(self.fileNameList)

    def createJson(self, output_path):
        for file_name in self.fileNameList:
            print(file_name)
            name_str = file_name[0:-5]
            cr = ConfigReader(self.dirPath+os.sep+file_name, name_str)
            cr.outputJson(output_path)


cm = ConfigManager(input_dir_path)
cm.createJson(output_dir_path)
